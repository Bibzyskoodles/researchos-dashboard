"""
XLSForm parsing (Bible 8.7: setup should feel like uploading a file).
Reads the `survey` sheet of a standard XLSForm workbook and derives
questionnaire_items — question text, required flags, and skip logic
(`relevant` expressions) — without the manager re-specifying anything.

Deliberately a thin openpyxl reader rather than full pyxform compilation:
we need the compliance-relevant fields, not a deployable form.
"""
import io
from dataclasses import dataclass

from openpyxl import load_workbook

# Rows that aren't questions an enumerator asks aloud.
_NON_QUESTION_TYPES = {
    "begin group", "end group", "begin_group", "end_group",
    "begin repeat", "end repeat", "begin_repeat", "end_repeat",
    "calculate", "note", "start", "end", "today", "deviceid",
    "phonenumber", "audio", "image", "geopoint", "geotrace", "geoshape",
}


@dataclass
class ParsedQuestion:
    question_key: str
    question_text: str
    is_required: bool
    skip_logic: dict | None
    sort_order: int
    # 'text' | 'numeric' | 'select_one' | 'select_multiple' — drives the
    # answer control in both capture surfaces. choices only for selects.
    question_type: str = "text"
    choices: list | None = None


def _parse_choices_sheet(wb) -> dict[str, list[dict]]:
    """XLSForm `choices` sheet → {list_name: [{name, label}, ...]}."""
    sheet_name = next((n for n in wb.sheetnames if n.strip().lower() == "choices"), None)
    if sheet_name is None:
        return {}
    ws = wb[sheet_name]
    rows = ws.iter_rows(values_only=True)
    try:
        header = [str(c).strip().lower() if c is not None else "" for c in next(rows)]
    except StopIteration:
        return {}

    def col(*names: str) -> int | None:
        for n in names:
            if n in header:
                return header.index(n)
        for i, h in enumerate(header):
            if any(h.startswith(n + "::") for n in names):
                return i
        return None

    c_list, c_name, c_label = col("list_name", "list name"), col("name"), col("label")
    if c_list is None or c_name is None:
        return {}
    out: dict[str, list[dict]] = {}
    for row in rows:
        def cell(idx):
            if idx is None or idx >= len(row) or row[idx] is None:
                return ""
            return str(row[idx]).strip()
        list_name, name = cell(c_list), cell(c_name)
        if not list_name or not name:
            continue
        out.setdefault(list_name, []).append({"name": name, "label": cell(c_label) or name})
    return out


def parse_xlsform(data: bytes) -> list[ParsedQuestion]:
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    sheet_name = next((n for n in wb.sheetnames if n.strip().lower() == "survey"), None)
    if sheet_name is None:
        raise ValueError("Not an XLSForm: no 'survey' sheet found.")
    choice_lists = _parse_choices_sheet(wb)
    ws = wb[sheet_name]

    rows = ws.iter_rows(values_only=True)
    try:
        header = [str(c).strip().lower() if c is not None else "" for c in next(rows)]
    except StopIteration:
        raise ValueError("The survey sheet is empty.")

    def col(*names: str) -> int | None:
        for n in names:
            if n in header:
                return header.index(n)
        # label columns are often localized: label::English (en) etc.
        for i, h in enumerate(header):
            if any(h.startswith(n + "::") for n in names):
                return i
        return None

    c_type, c_name = col("type"), col("name")
    c_label, c_required, c_relevant = col("label"), col("required"), col("relevant")
    if c_type is None or c_name is None:
        raise ValueError("Not an XLSForm: survey sheet needs 'type' and 'name' columns.")

    out: list[ParsedQuestion] = []
    order = 0
    for row in rows:
        def cell(idx: int | None) -> str:
            if idx is None or idx >= len(row) or row[idx] is None:
                return ""
            return str(row[idx]).strip()

        qtype = cell(c_type).lower()
        name = cell(c_name)
        if not qtype or not name or qtype in _NON_QUESTION_TYPES:
            continue
        order += 1
        required_raw = cell(c_required).lower()
        relevant = cell(c_relevant)

        question_type, choices = "text", None
        parts = qtype.split()
        if parts[0] in ("select_one", "select_multiple") and len(parts) > 1:
            question_type = parts[0]
            choices = choice_lists.get(parts[1]) or None
        elif parts[0] in ("integer", "decimal", "range"):
            question_type = "numeric"

        out.append(ParsedQuestion(
            question_key=name,
            question_text=cell(c_label) or name,
            is_required=required_raw in ("yes", "true", "true()", "1"),
            skip_logic={"relevant": relevant} if relevant else None,
            sort_order=order,
            question_type=question_type,
            choices=choices,
        ))
    if not out:
        raise ValueError("No askable questions found in the survey sheet.")
    return out
