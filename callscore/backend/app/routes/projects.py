"""Project setup: import questionnaire, assign enumerators.
See docs/ARCHITECTURE_BIBLE.md Part 8.7 - setup should feel like uploading
a file, not configuring software.

Reconciled (docs/RECONCILIATION.md): projects themselves live in
fieldscore-backend (Supabase `projects`, TEXT PROJ-… ids) — this service
does NOT create projects. It only attaches Call-mode artifacts
(questionnaire_items) to an existing FieldScore project id.
"""
from fastapi import APIRouter, Depends, HTTPException, UploadFile
from pydantic import BaseModel
from sqlalchemy.orm import Session

from app import models
from app.agents.orchestrator import TIER_0
from app.db import get_db
from app.services.xlsform import parse_xlsform

router = APIRouter()


class CallConfigIn(BaseModel):
    consent_script: str
    consent_language: str = "en"
    jurisdiction: str | None = None
    # Speech-engine routing (provider names; null = language-aware default)
    stt_language: str | None = None
    stt_primary: str | None = None
    stt_verify: str | None = None


@router.put("/{project_id}/call-config")
def set_call_config(project_id: str, payload: CallConfigIn, db: Session = Depends(get_db)):
    """Consent script is project config, localized, displayed verbatim in
    the enumerator app so wording can't drift (Bible Part 7)."""
    if not payload.consent_script.strip():
        raise HTTPException(status_code=422, detail="consent_script cannot be empty.")
    cfg = db.get(models.CallProjectConfig, project_id)
    if cfg is None:
        cfg = models.CallProjectConfig(project_id=project_id, consent_script="")
        db.add(cfg)
    cfg.consent_script = payload.consent_script.strip()
    cfg.consent_language = payload.consent_language
    cfg.jurisdiction = payload.jurisdiction
    cfg.stt_language = payload.stt_language
    cfg.stt_primary = payload.stt_primary
    cfg.stt_verify = payload.stt_verify
    db.commit()
    from app.services import stt
    return {
        "project_id": project_id,
        "status": "saved",
        # Echo the effective engine order so a manager can see what their
        # choice actually resolves to with today's configured keys.
        "effective_stt_order": stt.resolve_order(
            language=payload.stt_language or payload.consent_language,
            primary=payload.stt_primary, verify=payload.stt_verify,
        ),
    }


@router.get("/{project_id}/call-config")
def get_call_config(project_id: str, db: Session = Depends(get_db)):
    cfg = db.get(models.CallProjectConfig, project_id)
    if cfg is None:
        raise HTTPException(status_code=404, detail="No call config for this project yet.")
    from app.services import stt
    return {
        "project_id": project_id,
        "consent_script": cfg.consent_script,
        "consent_language": cfg.consent_language,
        "jurisdiction": cfg.jurisdiction,
        "stt_language": cfg.stt_language,
        "stt_primary": cfg.stt_primary,
        "stt_verify": cfg.stt_verify,
        "effective_stt_order": stt.resolve_order(
            language=cfg.stt_language or cfg.consent_language,
            primary=cfg.stt_primary, verify=cfg.stt_verify,
        ),
    }


@router.post("/{project_id}/questionnaire")
async def import_questionnaire(project_id: str, file: UploadFile, db: Session = Depends(get_db)):
    """
    Parses an uploaded XLSForm and auto-derives questionnaire_items,
    including required flags and skip logic from the `relevant` column
    (Part 8.7) — a pure file upload, no manual re-specification of
    compliance rules. Re-importing replaces the project's items. Then runs
    the Tier 0 Questionnaire Design Agent (Bible 4A.2) — design-quality
    findings only; they never touch any interview's scorecard.
    """
    data = await file.read()
    try:
        questions = parse_xlsform(data)
    except ValueError as e:
        raise HTTPException(status_code=422, detail=str(e))

    db.query(models.QuestionnaireItem).filter(
        models.QuestionnaireItem.project_id == project_id
    ).delete()
    items = [
        models.QuestionnaireItem(
            project_id=project_id,
            question_key=q.question_key,
            question_text=q.question_text,
            is_required=q.is_required,
            skip_logic=q.skip_logic,
            sort_order=q.sort_order,
            question_type=q.question_type,
            choices=q.choices,
        )
        for q in questions
    ]
    db.add_all(items)
    db.commit()

    # Tier 0 review — setup-time only, never per-interview.
    design_findings = []
    for agent in TIER_0:
        try:
            findings = agent.run(project_id, {
                "questionnaire_items": [
                    {"question_key": q.question_key, "question_text": q.question_text,
                     "is_required": q.is_required, "skip_logic": q.skip_logic}
                    for q in questions
                ],
            })
            design_findings.extend(
                {"type": f.finding_type, "description": f.description, "confidence": f.confidence}
                for f in findings
            )
        except NotImplementedError:
            pass

    return {
        "project_id": project_id,
        "imported": len(items),
        "design_review": design_findings,
    }


class QuestionIn(BaseModel):
    question_key: str | None = None
    question_text: str
    question_type: str = "text"  # text | numeric | select_one | select_multiple
    is_required: bool = True
    choices: list | None = None  # [{name, label}] for the select types
    # {"role": "trap", "expected": str, "note": str} — supervisor-set
    # red herring; scoring raises trap_failed when answered otherwise.
    integrity: dict | None = None


class QuestionnaireIn(BaseModel):
    items: list[QuestionIn]


@router.put("/{project_id}/questionnaire")
def set_questionnaire(project_id: str, payload: QuestionnaireIn, db: Session = Depends(get_db)):
    """The no-XLSForm path: the dashboard's call-mode question editor
    saves the whole questionnaire as JSON. Full replace, same as re-import
    (Part 8.7 — setup should feel simple, not like configuring software)."""
    cleaned = [q for q in payload.items if q.question_text.strip()]
    if not cleaned:
        raise HTTPException(status_code=422, detail="At least one question is required.")

    db.query(models.QuestionnaireItem).filter(
        models.QuestionnaireItem.project_id == project_id
    ).delete()
    for order, q in enumerate(cleaned, start=1):
        qtype = q.question_type if q.question_type in (
            "text", "numeric", "select_one", "select_multiple") else "text"
        db.add(models.QuestionnaireItem(
            project_id=project_id,
            question_key=(q.question_key or "").strip() or f"q{order}",
            question_text=q.question_text.strip(),
            is_required=q.is_required,
            skip_logic=None,
            sort_order=order,
            question_type=qtype,
            choices=q.choices if qtype in ("select_one", "select_multiple") else None,
            integrity=q.integrity or None,
        ))
    db.commit()
    return {"project_id": project_id, "saved": len(cleaned)}


@router.get("/{project_id}/questionnaire")
def get_questionnaire(project_id: str, db: Session = Depends(get_db)):
    """Drives the enumerator app's Glance-Confirm rows and the Question
    Compliance agent — one source of truth for both."""
    items = (
        db.query(models.QuestionnaireItem)
        .filter(models.QuestionnaireItem.project_id == project_id)
        .order_by(models.QuestionnaireItem.sort_order)
        .all()
    )
    return {
        "project_id": project_id,
        "items": [
            {
                "question_key": i.question_key,
                "question_text": i.question_text,
                "is_required": i.is_required,
                "skip_logic": i.skip_logic,
                "sort_order": i.sort_order,
                "question_type": i.question_type,
                "choices": i.choices,
                "integrity": i.integrity,
            }
            for i in items
        ],
    }


@router.get("/{project_id}/analysis-export")
def analysis_export(project_id: str, db: Session = Depends(get_db)):
    """Verified (PASS) call interviews for InsightScore — the pull side of
    the handoff (RECONCILIATION.md: InsightScore is the consumer, Bible
    Part 11). Same payload the outbox push enqueues, so either transport
    yields identical analysis input: answers + full transcript + shared
    scoring vocabulary."""
    from app.services.insight_bridge import build_analysis_payload

    rows = (
        db.query(models.Submission)
        .filter(
            models.Submission.project_id == project_id,
            models.Submission.collection_mode == "call",
            models.Submission.verdict == "PASS",
        )
        .order_by(models.Submission.started_at.asc())
        .limit(2000)
        .all()
    )
    return {
        "project_id": project_id,
        "collection_mode": "call",
        "count": len(rows),
        "interviews": [build_analysis_payload(db, s) for s in rows],
    }


# ── Ada-assisted questionnaire creation ─────────────────────────────────
# Two more ways in besides typing and strict XLSForm: describe the study
# and Ada drafts it, or upload ANY spreadsheet/CSV and Ada reads the
# questions out of it. Both return items for the editor to review — the
# human saves; Ada never writes the questionnaire herself (AI drafts,
# humans approve).

_QUESTION_CONTRACT = (
    'Respond with JSON: {"items": [{"question_text": string, '
    '"question_type": "text"|"numeric"|"select_one"|"select_multiple", '
    '"is_required": boolean, '
    '"choices": [{"name": string, "label": string}] | null}]}. '
    "choices only for the select types. Questions must be neutral, "
    "single-purpose, and phrased exactly as an interviewer would ask them "
    "aloud. Never invent content that is not implied by the input."
)


def _clean_drafted_items(result: dict | None) -> list[dict]:
    if not result:
        return []
    out = []
    for q in result.get("items", []):
        text = str(q.get("question_text", "")).strip()
        if not text:
            continue
        qtype = q.get("question_type")
        if qtype not in ("text", "numeric", "select_one", "select_multiple"):
            qtype = "text"
        choices = q.get("choices") if qtype in ("select_one", "select_multiple") else None
        if choices is not None:
            choices = [
                {"name": str(c.get("name") or c.get("label", "")).strip().lower().replace(" ", "_"),
                 "label": str(c.get("label", "")).strip()}
                for c in choices if str(c.get("label", "")).strip()
            ] or None
        out.append({
            "question_text": text,
            "question_type": qtype,
            "is_required": bool(q.get("is_required", True)),
            "choices": choices,
        })
    return out[:60]


class DraftBrief(BaseModel):
    brief: str


@router.post("/{project_id}/questionnaire/draft")
def draft_questionnaire(project_id: str, payload: DraftBrief):
    """Ada drafts a questionnaire from a plain-language study brief."""
    from app.services import llm

    if not llm.available():
        raise HTTPException(status_code=503, detail="Ada's drafting needs the AI service configured.")
    brief = payload.brief.strip()
    if not brief:
        raise HTTPException(status_code=422, detail="Describe the study in a sentence or two first.")
    result = llm.judge(
        "You are an expert survey methodologist drafting a phone-interview "
        "questionnaire from a study brief. 6-12 questions unless the brief "
        "implies otherwise, ordered for natural conversation flow (easy "
        "opener first, sensitive topics later). " + _QUESTION_CONTRACT,
        f"STUDY BRIEF:\n{brief[:4000]}",
    )
    items = _clean_drafted_items(result)
    if not items:
        raise HTTPException(status_code=502, detail="Ada could not draft from that brief — try adding more detail.")
    return {"project_id": project_id, "items": items, "source": "ada_draft"}


@router.post("/{project_id}/questionnaire/parse-file")
async def parse_questionnaire_file(project_id: str, file: UploadFile):
    """Ada reads questions out of ANY spreadsheet or CSV. Strict XLSForms
    are parsed deterministically (no AI needed); anything else is read
    cell-by-cell and Ada extracts the questions for review."""
    data = await file.read()
    name = (file.filename or "").lower()

    # Deterministic first: a real XLSForm needs no AI.
    if name.endswith((".xlsx", ".xls", ".xlsm")):
        try:
            questions = parse_xlsform(data)
            return {
                "project_id": project_id,
                "source": "xlsform",
                "items": [
                    {"question_text": q.question_text, "question_type": q.question_type,
                     "is_required": q.is_required, "choices": q.choices}
                    for q in questions
                ],
            }
        except ValueError:
            pass  # not an XLSForm — fall through to Ada

    from app.services import llm

    if not llm.available():
        raise HTTPException(
            status_code=503,
            detail="This file isn't a standard form, and Ada's reading needs the AI service configured.",
        )

    # Flatten the file to text for Ada.
    text = ""
    if name.endswith((".xlsx", ".xls", ".xlsm")):
        import io as _io

        from openpyxl import load_workbook

        try:
            wb = load_workbook(_io.BytesIO(data), read_only=True, data_only=True)
        except Exception:
            raise HTTPException(status_code=422, detail="Could not open that file — is it a valid Excel file?")
        lines = []
        for sheet in wb.sheetnames[:5]:
            lines.append(f"--- sheet: {sheet} ---")
            for row in wb[sheet].iter_rows(values_only=True):
                cells = [str(c).strip() for c in row if c is not None and str(c).strip()]
                if cells:
                    lines.append(" | ".join(cells))
                if len(lines) > 600:
                    break
        text = "\n".join(lines)
    else:
        text = data.decode("utf-8-sig", errors="replace")

    if not text.strip():
        raise HTTPException(status_code=422, detail="The file looks empty.")

    result = llm.judge(
        "You extract interview questions from a research team's document "
        "(a spreadsheet or CSV dump). Identify the actual questions an "
        "interviewer asks a respondent — ignore headers, notes, metadata, "
        "logic columns and translations (keep one language, prefer "
        "English). Preserve the original order and wording. Detect answer "
        "options where present. " + _QUESTION_CONTRACT,
        f"DOCUMENT:\n{text[:24000]}",
    )
    items = _clean_drafted_items(result)
    if not items:
        raise HTTPException(status_code=422, detail="Ada could not find questions in that file.")
    return {"project_id": project_id, "items": items, "source": "ada_file"}


@router.post("/{project_id}/questionnaire/review")
def review_questionnaire(project_id: str, payload: QuestionnaireIn):
    """Ada reviews the questionnaire as drafted in the editor (unsaved
    items welcome): deterministic design heuristics (Tier 0 agent) plus,
    when the LLM is available, improvement suggestions and trap-question
    ideas. Advisory only — nothing is changed until the human saves."""
    items = [
        {"question_key": q.question_key or f"q{i+1}", "question_text": q.question_text,
         "is_required": q.is_required, "skip_logic": None}
        for i, q in enumerate(payload.items) if q.question_text.strip()
    ]
    if not items:
        raise HTTPException(status_code=422, detail="Nothing to review yet — add questions first.")

    findings = []
    for agent in TIER_0:
        try:
            findings.extend(
                {"type": f.finding_type, "description": f.description, "confidence": f.confidence}
                for f in agent.run(project_id, {"questionnaire_items": items})
            )
        except NotImplementedError:
            pass

    from app.services import llm
    suggestions = []
    if llm.available():
        import json as _json
        result = llm.judge(
            "You are an expert survey methodologist reviewing a phone-"
            "interview questionnaire. Suggest concrete improvements: "
            "clearer wording, better ordering, missing screener or "
            "demographic questions, and ONE integrity trap question idea "
            "(a plausible-sounding question about something fictitious, "
            "where any positive answer signals fabrication — state the "
            "expected honest answer). Respond with JSON: {\"suggestions\": "
            '[{"kind": "wording"|"ordering"|"missing"|"trap_idea", '
            '"suggestion": string, "question_key": string|null, '
            '"trap": {"question_text": string, "choices": [string], '
            '"expected": string} | null}]}. Max 6 suggestions, most '
            "valuable first. Never invent problems to seem thorough.",
            f"QUESTIONNAIRE:\n{_json.dumps(items, indent=1)[:8000]}",
        )
        for sug in (result or {}).get("suggestions", [])[:6]:
            if str(sug.get("suggestion", "")).strip():
                suggestions.append({
                    "kind": str(sug.get("kind", "wording")),
                    "suggestion": str(sug.get("suggestion"))[:400],
                    "question_key": sug.get("question_key"),
                    "trap": sug.get("trap") if isinstance(sug.get("trap"), dict) else None,
                })

    return {"project_id": project_id, "findings": findings, "suggestions": suggestions}
